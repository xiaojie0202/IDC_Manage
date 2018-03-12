import openpyxl
import datetime
from openpyxl.worksheet.cell_range import CellRange
from dc_info import models


def reduce(fun, arry, init=None):
    ret = init if init else arry.pop(0)
    for i in arry:
        ret = fun(ret, i)
    return ret

def handel_import_equipment(file, dcname=None, idcname=None):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active

    # 读取一列
    row = list(sheet.rows)
    split_row_in = []
    # 切出来个么对象的行
    for col in sheet.columns:
        start = 1
        for index, i in enumerate(col):
            if index == 0:
                continue
            if i.value:
                if index == 1:
                    continue
                split_row_in.append(row[start:index])
                start = index
        break
    equipment_obj_list = []
    for i in split_row_in:
        # 每一个I 都是一个机柜的对象
        equipment_obj = {}
        ipaddress_set = []
        portinfo_set = []
        for a in i: # 循环完1行
            ipaddress = {}
            portinfo = {}
            for b in a:
                cel_name = b.column
                value = b.value
                if value:
                    if cel_name == 'A':
                        equipment_obj['cabinet'] = value
                    if cel_name == 'B':
                        equipment_obj['equipment_type'] = value
                    if cel_name == 'C':
                        equipment_obj['manufacturers'] = value
                    if cel_name == 'D':
                        equipment_obj['model_num'] = value
                    if cel_name == 'E':
                        equipment_obj['serial_num'] = value
                    if cel_name == 'F':
                        equipment_obj['place_u'] = value
                    if cel_name == 'G':
                        equipment_obj['equipment_u'] = value
                    if cel_name == 'H':
                        equipment_obj['customer'] = value
                    if cel_name == 'I':
                        equipment_obj['up_date'] = value
                    if cel_name == 'J':
                        ipaddress['tags'] = value
                    if cel_name == 'K':
                        ipaddress['ipaddre'] = value
                    if cel_name == 'L':
                        ipaddress['netmask'] = value
                    if cel_name == 'M':
                        ipaddress['gateway'] = value
                    if cel_name == 'N':
                        portinfo['self_equipment_port'] = value
                    if cel_name == 'O':
                        portinfo['up_equipment'] = value
                    if cel_name == 'P':
                        portinfo['up_equipment_port'] = value
            if ipaddress:
                ipaddress_set.append(ipaddress)
            if portinfo:
                portinfo_set.append(portinfo)
        equipment_obj['ipaddrss_set'] = ipaddress_set
        equipment_obj['portinfo_set'] = portinfo_set
        equipment_obj_list.append(equipment_obj)
    return equipment_obj_list


def handel_import_inventory(file, dcname, idcname):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active
    info = {'status':True, 'erro':None}
    inventory_objlist = []
    for item, excel_row in enumerate(sheet.rows):
        if item == 0:
            continue
        inventory_obj = {}
        for cell in excel_row:
            cell_num = cell.column
            value = cell.value
            if cell_num == 'A':
                inventory_obj['place'] = value
            if cell_num == 'B':
                inventory_obj['name'] = value
            if cell_num == 'C':
                inventory_obj['name_num'] = value
            if cell_num == 'D':
                inventory_obj['sn'] = value
            if cell_num == 'E':
                inventory_obj['count'] = value
            if cell_num == 'F':
                try:
                    inventory_obj['customer'] = models.Customer.objects.get(name=value)
                except Exception as e:
                    info['erro'] = '%s客户不存在' % value
                    break
            if cell_num == 'G':
                inventory_obj['node'] = value
            if cell_num == 'H':
                inventory_obj['post_number'] = value
            if cell_num == 'I':
                if value == '完好':
                    inventory_obj['status'] = 1
                elif value == '损坏':
                    inventory_obj['status'] = 0
                else:
                    pass
        inventory_obj['idc'] = models.DcInfo.objects.get(name=dcname).idcinfo_set.get(name=idcname)
        inventory_objlist.append(inventory_obj)
    info['inventory_list'] = inventory_objlist
    return info


def handel_import_cabinet(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    info = {'status': True, "erro": None}
    cabinet_obj_list = []
    for item, excel_row in enumerate(sheet.rows):
        if item == 0:
            continue
        cabinet_obj = {}
        dcname = None
        for cell in excel_row:
            cell_num = cell.column
            value = cell.value
            if cell_num == "A":
                dcname = value
            if cell_num == "B":
                cabinet_obj['idc'] = models.IDCInfo.objects.get(dc__name=dcname, name=value)
            if cell_num == "C":
                cabinet_obj['number'] = value
            if cell_num == "D":
                cabinet_obj['customer'] = models.Customer.objects.get(name=value)
            if cell_num == "E":
                if isinstance(value, datetime.datetime):
                    cabinet_obj['open_date'] = value
                else:
                    print(value, '不是日期类型')




if __name__ == '__main__':
    handel_import_cabinet("F:/3.xlsx")