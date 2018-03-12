from dc_info import models
from django.utils import timezone
import openpyxl
import datetime



def handel_cabinet(file):
    cabinet_num = []
    customer = []
    open_date = []
    cabinet_date_excel = openpyxl.load_workbook(file)  # 打开用户上传的excel文件
    myworksheet = cabinet_date_excel.active  # 文件活动工作表
    excel_row = myworksheet.max_row  # 获取工作表的最大行数
    excel_col = myworksheet.max_column  # 获取工作表的最大列数
    for col in range(1, excel_col + 1):
        for row in range(2, excel_row + 1):
            if col == 1:
                cabinet_num.append(myworksheet.cell(row=row, column=col).value)
            elif col == 2:
                customer.append(myworksheet.cell(row=row, column=col).value)
            elif col == 3:
                open_date.append(myworksheet.cell(row=row, column=col).value)
    cabinet_dict = {'cabinet_num': cabinet_num, 'customer': customer, 'open_date': open_date}
    return cabinet_dict

# 处理用户上传的批量导入机柜Excel表格
def handel_import_cabinet(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    info = {'status': True, "erro": None}
    cabinet_obj_list = []
    for item, excel_row in enumerate(sheet.rows):
        if item == 0:
            continue
        cabinet_obj = {}
        dcname = None
        for cell in excel_row:
            cell_num = cell.column
            value = cell.value
            if cell_num == "A":
                dcname = value
            if cell_num == "B":
                try:
                    cabinet_obj['idc'] = models.IDCInfo.objects.get(dc__name=dcname, name=value)
                except  Exception as e:
                    info['status'] = False
                    info['erro'] = "此机房不存在%s-%s,(%s)" % (dcname, value, e)
                    break
            if cell_num == "C":
                cabinet_obj['number'] = value
            if cell_num == "D":
                try:
                    cabinet_obj['customer'] = models.Customer.objects.get(name=value)
                except Exception as e:
                    info['status'] = False
                    info['erro'] = "%s此客户不存在,(%s)" % (value, e)
                    break
            if cell_num == "E":
                if isinstance(value, datetime.datetime):
                    cabinet_obj['open_date'] = value
                else:
                    info['status'] = False
                    info['erro'] = "%s不是日期类型" % value
                    break
        if info['status']:
            cabinet_obj_list.append(cabinet_obj)
        else:
            break
    info['cabinet_objlist'] = cabinet_obj_list
    return info

# 处理用户上传的批量导入设备excel表格
def handel_import_equipment(file, dcname, idcname):
    erro = {'status': True}
    cabinet_set = models.DcInfo.objects.get(name__contains=dcname).idcinfo_set.get(name__contains=idcname).cabinet_set
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active
    # 读取一列
    row = list(sheet.rows)
    split_row_in = []
    # 切出来个么对象的行
    for col in sheet.columns:
        start = 1
        for index, i in enumerate(col):
            if index == 0:
                continue
            if i.value:
                if index == 1:
                    continue
                split_row_in.append(row[start:index])
                start = index
        split_row_in.append(row[start:])
        break
    equipment_obj_list = []
    for i in split_row_in:
        # 每一个I 都是一个机柜的对象
        equipment_obj = {}
        ipaddress_set = []
        portinfo_set = []
        for a in i:  # 循环完1行
            ipaddress = {}
            portinfo = {}
            for b in a:
                cel_name = b.column
                try:
                    value = b.value.strip()
                except Exception as e:
                    value = b.value
                if value:
                    if cel_name == 'A':
                        try:
                            equipment_obj['cabinet'] = cabinet_set.get(number=value)
                        except Exception as e:
                            erro['erro'] = '%s机柜不存在,(%s)' % (value, e)
                            erro['status'] = False
                    if cel_name == 'B':
                        if value == '服务器设备':
                            equipment_obj['equipment_type'] = 1
                        elif value == '网络设备':
                            equipment_obj['equipment_type'] = 2
                        else:
                            erro['erro'] = '设备类型字段不正确'
                            erro['status'] = False
                    if cel_name == 'C':
                        equipment_obj['manufacturers'] = value
                    if cel_name == 'D':
                        equipment_obj['model_num'] = value
                    if cel_name == 'E':
                        equipment_obj['serial_num'] = value
                    if cel_name == 'F':
                        equipment_obj['place_u'] = value
                    if cel_name == 'G':
                        equipment_obj['equipment_u'] = value
                    if cel_name == 'H':
                        try:
                            customer = models.Customer.objects.get(name__contains=value)
                            equipment_obj['customer'] = customer
                        except Exception as e:
                            erro['erro'] = '%s客户不存在,(%s)' % (value, e)
                            erro['status'] = False
                    if cel_name == 'I':
                        try:
                            equipment_obj['up_date'] = timezone.datetime.date(value)
                        except Exception as e:
                            erro['erro'] = '%s日期格式不正确,(%s)' % (value, e)
                            erro['status'] = False
                    if cel_name == 'J':
                        ipaddress['tags'] = value
                    if cel_name == 'K':
                        ipaddress['ipaddre'] = value
                    if cel_name == 'L':
                        ipaddress['netmask'] = value
                    if cel_name == 'M':
                        ipaddress['gateway'] = value
                    if cel_name == 'N':
                        portinfo['self_equipment_port'] = value
                    if cel_name == 'O':
                        portinfo['up_equipment'] = value
                    if cel_name == 'P':
                        portinfo['up_equipment_port'] = value
            if ipaddress:
                ipaddress_set.append(ipaddress)
            if portinfo:
                portinfo_set.append(portinfo)
        equipment_obj['ipaddrss_set'] = ipaddress_set
        equipment_obj['portinfo_set'] = portinfo_set
        equipment_obj_list.append(equipment_obj)

    return equipment_obj_list, erro


# 处理用户批量导入库存信息Excel表格
def handel_import_inventory(file, dcname, idcname):
    wb = openpyxl.load_workbook(filename=file)
    sheet = wb.active
    info = {'status': True, 'erro': None}
    inventory_objlist = []
    for item, excel_row in enumerate(sheet.rows):
        if item == 0:
            continue
        inventory_obj = {}
        for cell in excel_row:
            cell_num = cell.column
            try:
                value = cell.value.strip()
            except Exception:
                value = cell.value
            if cell_num == 'A':
                inventory_obj['place'] = value
            if cell_num == 'B':
                inventory_obj['name'] = value
            if cell_num == 'C':
                inventory_obj['name_num'] = value
            if cell_num == 'D':
                inventory_obj['sn'] = value
            if cell_num == 'E':
                inventory_obj['count'] = value
            if cell_num == 'F':
                try:
                    inventory_obj['customer'] = models.Customer.objects.get(name=value)
                except Exception as e:
                    info['status'] = False
                    info['erro'] = '%s客户不存在' % value
                    break
            if cell_num == 'G':
                inventory_obj['node'] = value
            if cell_num == 'H':
                inventory_obj['post_number'] = value
            if cell_num == 'I':
                if value == '完好':
                    inventory_obj['status'] = 1
                elif value == '损坏':
                    inventory_obj['status'] = 0
                else:
                    info['status'] = False
                    info['erro'] = '%s 状态字段不正确' % value
                    break
        if info['status']:
            inventory_obj['idc'] = models.DcInfo.objects.get(name=dcname).idcinfo_set.get(name=idcname)
            try:
                inventory = models.Inventory.objects.create(**inventory_obj)
            except Exception as e:
                info['status'] = False
                info['erro'] = '当前资产已经存在(%s)' % inventory_obj
            else:
                inventory_objlist.append(inventory)
        else:
            break
    info['inventory_list'] = inventory_objlist
    return info